"""Export scraped results to JSON, CSV, or Excel."""

import json
import csv
import os
from pathlib import Path
from typing import List

from app.core.config import settings
from app.core.logging import get_logger
from app.models.db_models import ScrapedResult

logger = get_logger(__name__)


class ExportService:

    def __init__(self):
        self.export_dir = Path(settings.EXPORT_DIR)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _flatten(self, data: dict, prefix: str = "") -> dict:
        """Flatten nested dicts for CSV/Excel export."""
        flat = {}
        for k, v in data.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                flat.update(self._flatten(v, key))
            elif isinstance(v, list):
                flat[key] = json.dumps(v)
            else:
                flat[key] = v
        return flat

    def export_json(self, job_id: str, results: List[ScrapedResult]) -> str:
        output_path = self.export_dir / f"{job_id}.json"
        data = [
            {
                "url": r.url,
                "title": r.page_title,
                "status_code": r.status_code,
                "scraped_at": r.scraped_at.isoformat() if r.scraped_at else None,
                "extracted": r.extracted_data,
            }
            for r in results
        ]
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"Exported {len(results)} results to {output_path}")
        return str(output_path)

    def export_csv(self, job_id: str, results: List[ScrapedResult]) -> str:
        output_path = self.export_dir / f"{job_id}.csv"

        rows = []
        for r in results:
            base = {
                "url": r.url,
                "title": r.page_title,
                "status_code": r.status_code,
                "scraped_at": r.scraped_at.isoformat() if r.scraped_at else "",
            }
            if r.extracted_data:
                base.update(self._flatten(r.extracted_data))
            rows.append(base)

        if not rows:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                f.write("")
            return str(output_path)

        fieldnames = list(rows[0].keys())
        # Include any extra keys from later rows
        for row in rows[1:]:
            for k in row:
                if k not in fieldnames:
                    fieldnames.append(k)

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"Exported {len(results)} results to {output_path}")
        return str(output_path)

    def export_excel(self, job_id: str, results: List[ScrapedResult]) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        output_path = self.export_dir / f"{job_id}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Data"

        rows = []
        for r in results:
            base = {
                "URL": r.url,
                "Title": r.page_title,
                "Status": r.status_code,
                "Scraped At": r.scraped_at.isoformat() if r.scraped_at else "",
            }
            if r.extracted_data:
                base.update({k.title(): v for k, v in self._flatten(r.extracted_data).items()})
            rows.append(base)

        if rows:
            headers = list(rows[0].keys())
            # Style header row
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    val = row.get(header, "")
                    if isinstance(val, (dict, list)):
                        val = json.dumps(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(output_path)
        logger.info(f"Exported {len(results)} results to {output_path}")
        return str(output_path)


export_service = ExportService()
